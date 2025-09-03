#!/usr/bin/env python3
"""
Script d'analyse des heures disponibles par année et prix spot
Génère des tableaux Excel et graphiques avec les années en lignes
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

# Configuration
SEUILS_PRIX = [5, 10, 15, 20, 25, 30, 35, 40]  # €/MWh
DOSSIER_SORTIE = "heure_vs_puissance"

def creer_dossier_sortie():
    """Crée le dossier de sortie s'il n'existe pas"""
    if not os.path.exists(DOSSIER_SORTIE):
        os.makedirs(DOSSIER_SORTIE)
        print(f"Dossier créé : {DOSSIER_SORTIE}")

def charger_donnees():
    """Charge les données des prix spot"""
    try:
        df = pd.read_csv('donnees_prix_spot_processed_2020_2025.csv')
        print(f"Données chargées : {len(df)} lignes")
        print(f"Années disponibles : {sorted(df['Annee'].unique())}")
        return df
    except FileNotFoundError:
        print("Erreur : fichier 'donnees_prix_spot_processed_2020_2025.csv' non trouvé")
        return None

def calculer_heures_disponibles(df_annee, seuil_prix):
    """
    Calcule le nombre d'heures où le prix est ≤ au seuil
    """
    heures_disponibles = len(df_annee[df_annee['Prix'] <= seuil_prix])
    return heures_disponibles

def creer_tableau_global(df):
    """Crée le tableau des heures disponibles avec les années en lignes"""
    annees = sorted(df['Annee'].unique())
    
    # Créer le tableau de résultats
    resultats = []
    
    for annee in annees:
        df_annee = df[df['Annee'] == annee].copy()
        ligne = [int(annee)]  # Convertir en int pour un affichage plus propre
        
        for seuil in SEUILS_PRIX:
            heures = calculer_heures_disponibles(df_annee, seuil)
            ligne.append(heures)
        resultats.append(ligne)
    
    # Créer le DataFrame
    colonnes = ['Année'] + [f"{seuil}€/MWh" for seuil in SEUILS_PRIX]
    df_resultat = pd.DataFrame(resultats, columns=colonnes)
    
    return df_resultat

def formater_feuille_excel(ws):
    """Formate une feuille Excel avec styles"""
    # Style pour l'en-tête principal
    ws['A1'] = 'HEURES DISPONIBLES PAR ANNÉE ET SEUIL DE PRIX'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:I1')
    
    # Style pour les en-têtes de colonnes
    for col in range(1, 10):  # A à I
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Style pour les données
    nb_annees = len([cell for cell in ws['A'] if cell.value and str(cell.value).isdigit()])
    for row in range(4, 4 + nb_annees):  # Lignes de données dynamiques
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Coloration conditionnelle pour les valeurs
            if col > 1:  # Colonnes de données (pas la colonne année)
                valeur = cell.value
                if isinstance(valeur, (int, float)):
                    if valeur > 6000:  # Plus de 6000 heures (environ 70% de l'année)
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif valeur > 3000:  # Entre 3000 et 6000 heures
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:  # Moins de 3000 heures
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 15
    for col in range(2, 10):
        ws.column_dimensions[chr(64 + col)].width = 12

def sauvegarder_excel(df_resultat):
    """Sauvegarde le tableau dans un fichier Excel"""
    fichier_excel = os.path.join(DOSSIER_SORTIE, 'analyse_heures_disponibles_par_annee.xlsx')
    
    try:
        with pd.ExcelWriter(fichier_excel, engine='openpyxl') as writer:
            # Écrire les données
            df_resultat.to_excel(writer, sheet_name='Heures_par_annee', 
                               startrow=2, index=False)
            
            # Formater la feuille
            ws = writer.sheets['Heures_par_annee']
            formater_feuille_excel(ws)
        
        print(f"Fichier Excel sauvegardé : {fichier_excel}")
        
    except PermissionError:
        print(f"ERREUR : Impossible d'écrire le fichier {fichier_excel}")
        print("Le fichier est probablement ouvert dans Excel.")
        print("Solutions :")
        print("1. Fermez le fichier Excel s'il est ouvert")
        print("2. Ou renommez le fichier existant")
        
        # Essayer avec un nom alternatif
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fichier_alternatif = os.path.join(DOSSIER_SORTIE, f'analyse_heures_disponibles_par_annee_{timestamp}.xlsx')
        
        try:
            with pd.ExcelWriter(fichier_alternatif, engine='openpyxl') as writer:
                df_resultat.to_excel(writer, sheet_name='Heures_par_annee', 
                                   startrow=2, index=False)
                ws = writer.sheets['Heures_par_annee']
                formater_feuille_excel(ws)
            
            print(f"Fichier alternatif sauvegardé : {fichier_alternatif}")
            
        except Exception as e:
            print(f"Erreur lors de la sauvegarde alternative : {e}")
            # Sauvegarder au moins en CSV
            fichier_csv = os.path.join(DOSSIER_SORTIE, f'analyse_heures_disponibles_par_annee_{timestamp}.csv')
            df_resultat.to_csv(fichier_csv, index=False)
            print(f"Données sauvegardées en CSV : {fichier_csv}")
    
    except Exception as e:
        print(f"Erreur inattendue lors de la sauvegarde Excel : {e}")
        # Sauvegarder en CSV comme solution de secours
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fichier_csv = os.path.join(DOSSIER_SORTIE, f'analyse_heures_disponibles_par_annee_{timestamp}.csv')
        df_resultat.to_csv(fichier_csv, index=False)
        print(f"Données sauvegardées en CSV : {fichier_csv}")

def creer_graphique_principal(df_resultat):
    """Crée un graphique principal avec les années en axe X"""
    # Préparer les données pour le graphique
    df_graph = df_resultat.set_index('Année')
    
    # Créer le graphique
    plt.figure(figsize=(14, 8))
    
    # Graphique en barres groupées
    annees = df_resultat['Année'].values
    x = np.arange(len(annees))
    width = 0.1
    
    colors = plt.cm.viridis(np.linspace(0, 1, len(SEUILS_PRIX)))
    
    for i, seuil in enumerate(SEUILS_PRIX):
        colonne = f"{seuil}€/MWh"
        valeurs = df_graph[colonne].values
        plt.bar(x + i * width, valeurs, width, label=colonne, color=colors[i])
    
    # Personnalisation du graphique
    plt.title('Heures disponibles par année et seuil de prix', 
              fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Année', fontsize=12, fontweight='bold')
    plt.ylabel('Heures disponibles', fontsize=12, fontweight='bold')
    
    # Configuration des axes
    plt.xticks(x + width * 3.5, annees)
    plt.ylim(0, 9000)
    
    # Grille
    plt.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Légende
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', 
               title='Seuil de prix', title_fontsize=12)
    
    # Ajustement de la mise en page
    plt.tight_layout()
    
    # Sauvegarder
    fichier_graph = os.path.join(DOSSIER_SORTIE, 'graphique_heures_disponibles_par_annee.png')
    plt.savefig(fichier_graph, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphique principal sauvegardé : {fichier_graph}")

def creer_graphique_heatmap(df_resultat):
    """Crée une heatmap des heures disponibles"""
    plt.figure(figsize=(12, 8))
    
    # Préparer les données pour la heatmap
    df_heatmap = df_resultat.set_index('Année')
    
    # Créer la heatmap
    sns.heatmap(df_heatmap, 
               annot=True, 
               fmt='d',
               cmap='RdYlGn',
               cbar_kws={'label': 'Heures disponibles'})
    
    plt.title('Heatmap des heures disponibles par année et seuil de prix', 
             fontsize=14, fontweight='bold')
    plt.xlabel('Seuil de prix', fontsize=12)
    plt.ylabel('Année', fontsize=12)
    
    # Sauvegarder
    fichier_heatmap = os.path.join(DOSSIER_SORTIE, 'heatmap_heures_disponibles.png')
    plt.tight_layout()
    plt.savefig(fichier_heatmap, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Heatmap sauvegardée : {fichier_heatmap}")

def creer_graphiques_evolution(df_resultat):
    """Crée un graphique d'évolution avec tous les seuils de prix en même temps"""
    plt.figure(figsize=(14, 8))
    
    # Couleurs pour chaque seuil
    colors = plt.cm.tab10(np.linspace(0, 1, len(SEUILS_PRIX)))
    
    # Tracer une courbe pour chaque seuil
    for i, seuil in enumerate(SEUILS_PRIX):
        colonne = f"{seuil}€/MWh"
        
        plt.plot(df_resultat['Année'], df_resultat[colonne], 
                marker='o', linewidth=2.5, markersize=8, 
                label=f'{seuil}€/MWh', color=colors[i])
    
    # Personnalisation
    plt.title('Évolution des heures disponibles par seuil de prix', 
             fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Année', fontsize=12, fontweight='bold')
    plt.ylabel('Heures disponibles', fontsize=12, fontweight='bold')
    
    # Configuration des axes
    plt.xticks(df_resultat['Année'])
    
    # Déterminer la limite Y maximale
    max_valeur = 0
    for seuil in SEUILS_PRIX:
        colonne = f"{seuil}€/MWh"
        max_valeur = max(max_valeur, max(df_resultat[colonne]))
    plt.ylim(0, max_valeur * 1.1)
    
    # Grille
    plt.grid(True, alpha=0.3, linestyle='--')
    
    # Légende
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left', 
               title='Seuil de prix', title_fontsize=12, fontsize=10)
    
    plt.tight_layout()
    
    # Sauvegarder
    fichier = os.path.join(DOSSIER_SORTIE, 'evolution_heures_tous_seuils.png')
    plt.savefig(fichier, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphique d'évolution combiné sauvegardé : {fichier}")

def analyser_donnees():
    """Fonction principale d'analyse"""
    print("=== ANALYSE DES HEURES DISPONIBLES PAR ANNÉE ET PRIX ===")
    print(f"Démarrage de l'analyse : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Créer le dossier de sortie
    creer_dossier_sortie()
    
    # Charger les données
    df = charger_donnees()
    if df is None:
        return
    
    # Créer le tableau global avec les années en lignes
    print("\n--- Création du tableau principal ---")
    df_resultat = creer_tableau_global(df)
    
    # Afficher un aperçu
    print("Aperçu des résultats :")
    print(df_resultat)
    
    # Sauvegarder le fichier Excel
    sauvegarder_excel(df_resultat)
    
    # Créer les graphiques
    print("\n--- Création des graphiques ---")
    creer_graphique_principal(df_resultat)
    creer_graphique_heatmap(df_resultat)
    creer_graphiques_evolution(df_resultat)
    
    print(f"\n=== ANALYSE TERMINÉE ===")
    print(f"Résultats disponibles dans le dossier : {DOSSIER_SORTIE}")

if __name__ == "__main__":
    analyser_donnees() 