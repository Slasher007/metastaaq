#!/usr/bin/env python3
"""
Script d'analyse des heures disponibles par année et mois pour différents seuils de prix
Génère des tableaux Excel et graphiques avec comparaison de plusieurs seuils
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

# Configuration
SEUILS_PRIX = [5, 10, 15, 20, 25, 30]  # €/MWh - Plusieurs seuils à analyser
SEUIL_PRIX_PRINCIPAL = 15  # Seuil principal pour les analyses détaillées
DOSSIER_SORTIE = "analyse_mensuel"
MOIS_NOMS = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 
            'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']

def creer_dossier_sortie():
    """Crée le dossier de sortie s'il n'existe pas"""
    if not os.path.exists(DOSSIER_SORTIE):
        os.makedirs(DOSSIER_SORTIE)
        print(f"Dossier créé : {DOSSIER_SORTIE}")

def charger_donnees():
    """Charge les données des prix spot"""
    try:
        df = pd.read_csv('donnees_prix_spot_processed_2020_2025.csv')
        
        # Convertir la colonne Date en datetime si elle ne l'est pas déjà
        if 'Date' in df.columns:
            df['Date'] = pd.to_datetime(df['Date'])
            df['Mois'] = df['Date'].dt.month
        elif 'Mois' not in df.columns:
            print("Attention : Colonnes disponibles :", df.columns.tolist())
            # Essayer de créer la colonne Mois à partir d'autres colonnes
            if 'date' in df.columns:
                df['Date'] = pd.to_datetime(df['date'])
                df['Mois'] = df['Date'].dt.month
            else:
                print("Erreur : Impossible de déterminer le mois")
                return None
        
        print(f"Données chargées : {len(df)} lignes")
        print(f"Années disponibles : {sorted(df['Annee'].unique())}")
        print(f"Mois disponibles : {sorted(df['Mois'].unique())}")
        return df
    except FileNotFoundError:
        print("Erreur : fichier 'donnees_prix_spot_processed_2020_2025.csv' non trouvé")
        return None

def calculer_heures_par_mois(df_annee_mois, seuil_prix):
    """
    Calcule le nombre d'heures où le prix est ≤ au seuil pour un mois donné
    """
    heures_disponibles = len(df_annee_mois[df_annee_mois['Prix'] <= seuil_prix])
    return heures_disponibles

def creer_tableau_mensuel(df, seuil_prix):
    """Crée le tableau des heures disponibles avec les années en lignes et mois en colonnes"""
    annees = sorted(df['Annee'].unique())
    
    # Créer le tableau de résultats
    resultats = []
    
    for annee in annees:
        df_annee = df[df['Annee'] == annee].copy()
        ligne = [int(annee)]  # Année en première colonne
        
        # Pour chaque mois
        for mois in range(1, 13):
            df_mois = df_annee[df_annee['Mois'] == mois]
            if len(df_mois) > 0:
                heures = calculer_heures_par_mois(df_mois, seuil_prix)
            else:
                heures = 0  # Aucune donnée pour ce mois
            ligne.append(heures)
        
        resultats.append(ligne)
    
    # Créer le DataFrame
    colonnes = ['Année'] + MOIS_NOMS
    df_resultat = pd.DataFrame(resultats, columns=colonnes)
    
    return df_resultat

def creer_tableaux_tous_seuils(df):
    """Crée les tableaux pour tous les seuils de prix"""
    tableaux = {}
    for seuil in SEUILS_PRIX:
        print(f"Calcul pour seuil {seuil}€/MWh...")
        tableaux[seuil] = creer_tableau_mensuel(df, seuil)
    return tableaux

def creer_graphique_comparatif_seuils(tableaux_seuils):
    """Crée un graphique comparatif avec tous les seuils de prix"""
    plt.figure(figsize=(18, 12))
    
    # Couleurs distinctes pour chaque seuil
    colors = plt.cm.tab10(np.linspace(0, 1, len(SEUILS_PRIX)))
    
    mois_num = range(1, 13)
    
    # Calculer les moyennes par mois pour chaque seuil (toutes années confondues)
    for i, seuil in enumerate(SEUILS_PRIX):
        df_seuil = tableaux_seuils[seuil]
        moyennes_mensuelles = df_seuil[MOIS_NOMS].mean(axis=0)
        
        plt.plot(mois_num, moyennes_mensuelles, 
                marker='o', linewidth=3, markersize=8, 
                label=f'{seuil}€/MWh', color=colors[i],
                alpha=0.8)
    
    # Personnalisation du graphique
    plt.title('Comparaison des heures disponibles par mois selon différents seuils de prix\n(Moyenne sur toutes les années)', 
              fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Mois', fontsize=14, fontweight='bold')
    plt.ylabel('Heures disponibles (moyenne)', fontsize=14, fontweight='bold')
    
    # Configuration des axes
    plt.xticks(mois_num, MOIS_NOMS, fontsize=12)
    plt.yticks(fontsize=12)
    
    # Trouver la valeur max pour ajuster l'axe Y
    max_val = max([tableaux_seuils[seuil][MOIS_NOMS].mean(axis=0).max() for seuil in SEUILS_PRIX])
    plt.ylim(0, max_val * 1.1)
    
    # Grille
    plt.grid(True, alpha=0.3, linestyle='--')
    
    # Légende
    plt.legend(title='Seuil de prix', bbox_to_anchor=(1.05, 1), loc='upper left', 
               fontsize=12, title_fontsize=14)
    
    # Ajustement de la mise en page
    plt.tight_layout()
    
    # Sauvegarder
    fichier_graph = os.path.join(DOSSIER_SORTIE, 'comparaison_seuils_prix.png')
    plt.savefig(fichier_graph, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphique comparatif sauvegardé : {fichier_graph}")

def creer_graphique_evolution_annuelle_seuils(tableaux_seuils):
    """Crée un graphique montrant l'évolution du total annuel pour chaque seuil"""
    plt.figure(figsize=(16, 10))
    
    colors = plt.cm.tab10(np.linspace(0, 1, len(SEUILS_PRIX)))
    
    # Pour chaque seuil, calculer le total annuel
    for i, seuil in enumerate(SEUILS_PRIX):
        df_seuil = tableaux_seuils[seuil]
        totaux_annuels = df_seuil[MOIS_NOMS].sum(axis=1)
        
        plt.plot(df_seuil['Année'], totaux_annuels, 
                marker='o', linewidth=3, markersize=8, 
                label=f'{seuil}€/MWh', color=colors[i],
                alpha=0.8)
    
    plt.title('Évolution du total annuel d\'heures disponibles par seuil de prix', 
              fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Année', fontsize=14, fontweight='bold')
    plt.ylabel('Total d\'heures par an', fontsize=14, fontweight='bold')
    
    plt.grid(True, alpha=0.3, linestyle='--')
    plt.legend(title='Seuil de prix', bbox_to_anchor=(1.05, 1), loc='upper left',
               fontsize=12, title_fontsize=14)
    
    plt.tight_layout()
    
    fichier_graph = os.path.join(DOSSIER_SORTIE, 'evolution_annuelle_seuils.png')
    plt.savefig(fichier_graph, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphique évolution annuelle sauvegardé : {fichier_graph}")

def creer_heatmap_comparative(tableaux_seuils):
    """Crée une heatmap comparative pour tous les seuils"""
    fig, axes = plt.subplots(2, 3, figsize=(20, 12))
    axes = axes.flatten()
    
    for i, seuil in enumerate(SEUILS_PRIX):
        df_heatmap = tableaux_seuils[seuil].set_index('Année')[MOIS_NOMS]
        
        sns.heatmap(df_heatmap, 
                   annot=True, 
                   fmt='d',
                   cmap='RdYlGn',
                   ax=axes[i],
                   cbar_kws={'label': 'Heures'})
        
        axes[i].set_title(f'Seuil {seuil}€/MWh', fontsize=12, fontweight='bold')
        axes[i].set_xlabel('Mois')
        axes[i].set_ylabel('Année')
    
    plt.suptitle('Heatmaps comparatives des heures disponibles par seuil de prix', 
                 fontsize=16, fontweight='bold')
    plt.tight_layout()
    
    fichier_heatmap = os.path.join(DOSSIER_SORTIE, 'heatmaps_comparatives.png')
    plt.savefig(fichier_heatmap, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Heatmaps comparatives sauvegardées : {fichier_heatmap}")

def calculer_statistiques(df_resultat):
    """Calcule des statistiques sur le tableau"""
    stats = {}
    
    # Statistiques par année
    df_mois = df_resultat[MOIS_NOMS]
    stats['total_par_annee'] = df_mois.sum(axis=1)
    stats['moyenne_par_annee'] = df_mois.mean(axis=1)
    stats['min_par_annee'] = df_mois.min(axis=1)
    stats['max_par_annee'] = df_mois.max(axis=1)
    
    # Statistiques par mois (sur toutes les années)
    stats['moyenne_par_mois'] = df_mois.mean(axis=0)
    stats['min_par_mois'] = df_mois.min(axis=0)
    stats['max_par_mois'] = df_mois.max(axis=0)
    
    return stats

def formater_feuille_excel(ws):
    """Formate une feuille Excel avec styles"""
    # Style pour l'en-tête principal
    ws['A1'] = f'HEURES DISPONIBLES PAR ANNÉE ET MOIS (Prix ≤ {SEUIL_PRIX_PRINCIPAL}€/MWh)'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:M1')
    
    # Style pour les en-têtes de colonnes
    for col in range(1, 14):  # A à M (Année + 12 mois)
        cell = ws.cell(row=3, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Style pour les données
    nb_annees = len([cell for cell in ws['A'] if cell.value and str(cell.value).isdigit()])
    for row in range(4, 4 + nb_annees):  # Lignes de données
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Coloration conditionnelle pour les valeurs (sauf colonne année)
            if col > 1:  
                valeur = cell.value
                if isinstance(valeur, (int, float)):
                    # Seuils basés sur le nombre d'heures dans un mois
                    if valeur > 500:  # Plus de 500 heures (très bon)
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif valeur > 200:  # Entre 200 et 500 heures (moyen)
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    else:  # Moins de 200 heures (faible)
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 15
    for col in range(2, 14):
        ws.column_dimensions[chr(64 + col)].width = 10

def sauvegarder_excel(df_resultat, stats):
    """Sauvegarde le tableau et les statistiques dans un fichier Excel"""
    fichier_excel = os.path.join(DOSSIER_SORTIE, f'heures_mensuelles_{SEUIL_PRIX_PRINCIPAL}euros.xlsx')
    
    with pd.ExcelWriter(fichier_excel, engine='openpyxl') as writer:
        # Feuille principale
        df_resultat.to_excel(writer, sheet_name='Heures_mensuelles', 
                           startrow=2, index=False)
        
        # Feuille des statistiques
        df_stats = pd.DataFrame({
            'Année': df_resultat['Année'],
            'Total_annuel': stats['total_par_annee'],
            'Moyenne_mensuelle': stats['moyenne_par_annee'].round(1),
            'Min_mensuel': stats['min_par_annee'],
            'Max_mensuel': stats['max_par_annee']
        })
        df_stats.to_excel(writer, sheet_name='Statistiques_annuelles', index=False)
        
        # Statistiques par mois
        df_stats_mois = pd.DataFrame({
            'Mois': MOIS_NOMS,
            'Moyenne': stats['moyenne_par_mois'].round(1),
            'Minimum': stats['min_par_mois'],
            'Maximum': stats['max_par_mois']
        })
        df_stats_mois.to_excel(writer, sheet_name='Statistiques_mensuelles', index=False)
        
        # Formater la feuille principale
        ws = writer.sheets['Heures_mensuelles']
        formater_feuille_excel(ws)
    
    print(f"Fichier Excel sauvegardé : {fichier_excel}")

def creer_graphique_principal(df_resultat):
    """Crée un graphique principal avec les mois en axe X et courbes par année"""
    plt.figure(figsize=(16, 10))
    
    # Créer un graphique avec une courbe par année
    mois_num = range(1, 13)
    colors = plt.cm.tab10(np.linspace(0, 1, len(df_resultat)))
    
    for i, (_, row) in enumerate(df_resultat.iterrows()):
        annee = row['Année']
        valeurs = row[MOIS_NOMS].values
        plt.plot(mois_num, valeurs, marker='o', linewidth=2, markersize=6, 
                label=f'{int(annee)}', color=colors[i])
    
    # Personnalisation du graphique
    plt.title(f'Heures disponibles par mois (Prix ≤ {SEUIL_PRIX_PRINCIPAL}€/MWh)', 
              fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Mois', fontsize=12, fontweight='bold')
    plt.ylabel('Heures disponibles', fontsize=12, fontweight='bold')
    
    # Configuration des axes
    plt.xticks(mois_num, MOIS_NOMS, rotation=45)
    plt.ylim(0, max(df_resultat[MOIS_NOMS].max()) * 1.1)
    
    # Grille
    plt.grid(True, alpha=0.3)
    
    # Légende
    plt.legend(title='Année', bbox_to_anchor=(1.05, 1), loc='upper left')
    
    # Ajustement de la mise en page
    plt.tight_layout()
    
    # Sauvegarder
    fichier_graph = os.path.join(DOSSIER_SORTIE, f'evolution_mensuelle_{SEUIL_PRIX_PRINCIPAL}euros.png')
    plt.savefig(fichier_graph, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphique principal sauvegardé : {fichier_graph}")

def creer_heatmap(df_resultat):
    """Crée une heatmap des heures disponibles"""
    plt.figure(figsize=(14, 8))
    
    # Préparer les données pour la heatmap
    df_heatmap = df_resultat.set_index('Année')[MOIS_NOMS]
    
    # Créer la heatmap
    sns.heatmap(df_heatmap, 
               annot=True, 
               fmt='d',
               cmap='RdYlGn',
               cbar_kws={'label': 'Heures disponibles'})
    
    plt.title(f'Heatmap des heures disponibles par année et mois (Prix ≤ {SEUIL_PRIX_PRINCIPAL}€/MWh)', 
             fontsize=14, fontweight='bold')
    plt.xlabel('Mois', fontsize=12)
    plt.ylabel('Année', fontsize=12)
    
    # Sauvegarder
    fichier_heatmap = os.path.join(DOSSIER_SORTIE, f'heatmap_mensuelle_{SEUIL_PRIX_PRINCIPAL}euros.png')
    plt.tight_layout()
    plt.savefig(fichier_heatmap, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Heatmap sauvegardée : {fichier_heatmap}")

def creer_graphique_barres(df_resultat):
    """Crée un graphique en barres groupées"""
    plt.figure(figsize=(16, 10))
    
    # Préparer les données
    df_plot = df_resultat.set_index('Année')[MOIS_NOMS]
    
    # Graphique en barres
    ax = df_plot.T.plot(kind='bar', figsize=(16, 10), width=0.8)
    
    plt.title(f'Heures disponibles par mois et année (Prix ≤ {SEUIL_PRIX_PRINCIPAL}€/MWh)', 
              fontsize=16, fontweight='bold', pad=20)
    plt.xlabel('Mois', fontsize=12, fontweight='bold')
    plt.ylabel('Heures disponibles', fontsize=12, fontweight='bold')
    
    # Configuration
    plt.xticks(rotation=45)
    plt.grid(axis='y', alpha=0.3)
    plt.legend(title='Année', bbox_to_anchor=(1.05, 1), loc='upper left')
    
    plt.tight_layout()
    
    # Sauvegarder
    fichier_barres = os.path.join(DOSSIER_SORTIE, f'barres_mensuelles_{SEUIL_PRIX_PRINCIPAL}euros.png')
    plt.savefig(fichier_barres, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphique en barres sauvegardé : {fichier_barres}")

def creer_graphique_statistiques(stats, df_resultat):
    """Crée des graphiques de statistiques"""
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # Graphique 1: Total par année
    ax1.bar(df_resultat['Année'], stats['total_par_annee'])
    ax1.set_title('Total d\'heures par année')
    ax1.set_ylabel('Heures totales')
    ax1.grid(axis='y', alpha=0.3)
    
    # Graphique 2: Moyenne par mois (toutes années confondues)
    ax2.bar(MOIS_NOMS, stats['moyenne_par_mois'])
    ax2.set_title('Moyenne par mois (toutes années)')
    ax2.set_ylabel('Heures moyennes')
    ax2.tick_params(axis='x', rotation=45)
    ax2.grid(axis='y', alpha=0.3)
    
    # Graphique 3: Évolution de la moyenne mensuelle par année
    ax3.plot(df_resultat['Année'], stats['moyenne_par_annee'], marker='o', linewidth=2)
    ax3.set_title('Évolution de la moyenne mensuelle')
    ax3.set_ylabel('Heures moyennes par mois')
    ax3.grid(True, alpha=0.3)
    
    # Graphique 4: Écart (max - min) par année
    ecarts = stats['max_par_annee'] - stats['min_par_annee']
    ax4.bar(df_resultat['Année'], ecarts)
    ax4.set_title('Écart mensuel (max - min) par année')
    ax4.set_ylabel('Écart (heures)')
    ax4.grid(axis='y', alpha=0.3)
    
    plt.suptitle(f'Statistiques des heures disponibles (Prix ≤ {SEUIL_PRIX_PRINCIPAL}€/MWh)', 
                 fontsize=16, fontweight='bold')
    plt.tight_layout()
    
    # Sauvegarder
    fichier_stats = os.path.join(DOSSIER_SORTIE, f'statistiques_{SEUIL_PRIX_PRINCIPAL}euros.png')
    plt.savefig(fichier_stats, dpi=300, bbox_inches='tight', facecolor='white')
    plt.close()
    
    print(f"Graphiques de statistiques sauvegardés : {fichier_stats}")

def analyser_donnees_mensuelles():
    """Fonction principale d'analyse"""
    print("=== ANALYSE DES HEURES DISPONIBLES PAR ANNÉE ET MOIS ===")
    print(f"Seuils de prix analysés : {SEUILS_PRIX} €/MWh")
    print(f"Seuil principal détaillé : ≤ {SEUIL_PRIX_PRINCIPAL}€/MWh")
    print(f"Démarrage de l'analyse : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Créer le dossier de sortie
    creer_dossier_sortie()
    
    # Charger les données
    df = charger_donnees()
    if df is None:
        return
    
    # Créer les tableaux pour tous les seuils
    print("\n--- Création des tableaux pour tous les seuils ---")
    tableaux_seuils = creer_tableaux_tous_seuils(df)
    
    # Tableau principal pour le seuil de référence
    df_resultat = tableaux_seuils[SEUIL_PRIX_PRINCIPAL]
    
    # Calculer les statistiques pour le seuil principal
    stats = calculer_statistiques(df_resultat)
    
    # Afficher un aperçu
    print("Aperçu des résultats (seuil principal) :")
    print(df_resultat)
    print(f"\nTotal d'heures par année : {stats['total_par_annee'].tolist()}")
    
    # Sauvegarder le fichier Excel pour le seuil principal
    sauvegarder_excel(df_resultat, stats)
    
    # Créer les nouveaux graphiques comparatifs
    print("\n--- Création des graphiques comparatifs ---")
    creer_graphique_comparatif_seuils(tableaux_seuils)
    creer_graphique_evolution_annuelle_seuils(tableaux_seuils)
    creer_heatmap_comparative(tableaux_seuils)
    
    # Créer les graphiques détaillés pour le seuil principal
    print("\n--- Création des graphiques détaillés (seuil principal) ---")
    creer_graphique_principal(df_resultat)
    creer_heatmap(df_resultat)
    creer_graphique_barres(df_resultat)
    creer_graphique_statistiques(stats, df_resultat)
    
    print(f"\n=== ANALYSE TERMINÉE ===")
    print(f"Résultats disponibles dans le dossier : {DOSSIER_SORTIE}")
    print(f"Graphiques comparatifs créés pour les seuils : {SEUILS_PRIX} €/MWh")

if __name__ == "__main__":
    analyser_donnees_mensuelles() 